import io
import json
import logging
import re
import secrets
import uuid
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from hashlib import sha256
from typing import Any
from zoneinfo import ZoneInfo

from openpyxl import Workbook, load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.security import hash_password
from app.models.distributor import Distributor
from app.models.enums import ImportType, PointsLedgerStatus, PointsOperationType, TaskType, UserRole
from app.models.import_error_log import ImportErrorLog
from app.models.task import Task
from app.models.task_distributor import TaskDistributor
from app.models.user import User
from app.models.user_task_acceptance import UserTaskAcceptance
from app.services.email import EmailService
from app.services.points import PointsService

logger = logging.getLogger(__name__)

USER_IMPORT_HEADERS = [
    "Код участника",
    "Дистрибьютор",
    "Должность",
    "ФИО",
    "Email",
]
EMAIL_REGEX = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")
EXISTING_USER_EMAIL_ERROR = (
    "Email уже зарегистрирован в системе. "
    "Новый пользователь не создавался, существующий участник не изменялся."
)
MSK_TZ = ZoneInfo("Europe/Moscow")
ZERO = Decimal("0.00")
DECIMAL_STEP = Decimal("0.01")

SALES_TEMPLATE_HEADERS = [
    "Дистрибьютор",
    "Филиал",
    "Код клиента",
    "Клиент",
    "Адрес",
    "СВ ФИО",
    "СВ Код",
    "ТП ФИО",
    "ТП Код",
    "Дата",
    "Год",
    "Месяц",
    "Товар",
    "Кол-во кор",
    "Баллы ТП",
    "Баллы СВ",
]
SALES_HEADERS_ALIASES: dict[str, tuple[str, ...]] = {
    "Дистрибьютор": ("Дистрибьютор",),
    "Филиал": ("Филиал",),
    "Код клиента": ("Код клиента",),
    "Клиент": ("Клиент", "Название клиента"),
    "Адрес": ("Адрес", "Адрес клиента"),
    "СВ ФИО": ("СВ ФИО", "Супервайзер ФИО"),
    "СВ Код": ("СВ Код", "Супервайзер Код"),
    "ТП ФИО": ("ТП ФИО",),
    "ТП Код": ("ТП Код",),
    "Дата": ("Дата", "Дата документа"),
    "Год": ("Год",),
    "Месяц": ("Месяц",),
    "Товар": ("Товар", "Название товара"),
    "Кол-во кор": ("Кол-во кор", "Количество, кор"),
    "Баллы ТП": ("Баллы ТП", "Кол-во начисленных баллов ТП"),
    "Баллы СВ": ("Баллы СВ", "Кол-во начисленных баллов СВ"),
}
MONTH_NAME_TO_NUMBER = {
    "январь": 1,
    "февраль": 2,
    "март": 3,
    "апрель": 4,
    "май": 5,
    "июнь": 6,
    "июль": 7,
    "август": 8,
    "сентябрь": 9,
    "октябрь": 10,
    "ноябрь": 11,
    "декабрь": 12,
}
SALES_ROLE_LABELS = {"tp": "ТП", "sv": "СВ"}
SALES_IMPORT_COMMENT_PREFIX = "import_sales_row:"


def build_sales_import_comment(
    payload: dict[str, str],
    *,
    role: str,
    document_date: date,
    period_month: str,
) -> str:
    """Сохраняет значения ячеек Excel без склеивания через запятые."""
    row = {header: payload.get(header, "") for header in SALES_TEMPLATE_HEADERS}
    row["__role"] = SALES_ROLE_LABELS.get(role, role)
    row["__document_date"] = document_date.isoformat()
    row["__period_month"] = period_month
    return SALES_IMPORT_COMMENT_PREFIX + json.dumps(row, ensure_ascii=False)


def _clean_cell(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _normalize_email(value: str) -> str:
    return value.strip().lower()


def _is_valid_email(value: str) -> bool:
    return bool(EMAIL_REGEX.match(value))


def _temporary_password() -> str:
    return secrets.token_urlsafe(8)


def _normalize_decimal(value: Decimal) -> Decimal:
    return value.quantize(DECIMAL_STEP)


def _parse_decimal(value: Any) -> Decimal:
    if value is None:
        return ZERO
    if isinstance(value, Decimal):
        return _normalize_decimal(value)
    if isinstance(value, (int, float)):
        return _normalize_decimal(Decimal(str(value)))
    text = _clean_cell(value).replace(" ", "").replace(",", ".")
    if not text:
        return ZERO
    try:
        return _normalize_decimal(Decimal(text))
    except InvalidOperation as exc:
        raise ValueError(f"Некорректное число: {value}") from exc


def _parse_year(value: Any) -> int | None:
    text = _clean_cell(value)
    if not text:
        return None
    if not text.isdigit() or len(text) != 4:
        raise ValueError(f"Некорректный год: {value}")
    return int(text)


def _parse_month(value: Any) -> int | None:
    text = _clean_cell(value)
    if not text:
        return None
    if text.isdigit():
        month_num = int(text)
        if 1 <= month_num <= 12:
            return month_num
        raise ValueError(f"Некорректный месяц: {value}")
    normalized = text.lower()
    month_num = MONTH_NAME_TO_NUMBER.get(normalized)
    if month_num is None:
        raise ValueError(f"Некорректный месяц: {value}")
    return month_num


def _parse_document_date(value: Any) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = _clean_cell(value)
    if not text:
        raise ValueError("Не заполнена «Дата»")
    for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"Некорректная дата документа: {value}")


def _build_import_key(parts: list[str]) -> str:
    raw = "|".join(parts)
    return sha256(raw.encode("utf-8")).hexdigest()


def _consent_deadline(period_month: str) -> datetime:
    year, month = period_month.split("-")
    year_num = int(year)
    month_num = int(month)
    if month_num == 12:
        deadline_year = year_num + 1
        deadline_month = 1
    else:
        deadline_year = year_num
        deadline_month = month_num + 1
    return datetime(deadline_year, deadline_month, 10, 23, 59, 59, tzinfo=MSK_TZ)


class ImportsService:
    def __init__(self, db: AsyncSession) -> None:
        self.db = db
        self.points_service = PointsService(db)

    async def build_users_template(self) -> bytes:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "users"
        sheet.append(USER_IMPORT_HEADERS)
        sheet.append(["ТП-001", "ООО Торг-Сервис", "ТП", "Иванов Иван Иванович", "user@example.ru"])
        sheet.append(["СВ-001", "ООО Торг-Сервис", "СВ", "Петров Петр Петрович", "sv@example.ru"])

        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        return output.read()

    async def build_sales_template(self) -> bytes:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "sales"
        sheet.append(SALES_TEMPLATE_HEADERS)
        sheet.append(
            [
                "ООО Торг-Сервис",
                "Красноярск",
                "57712186",
                "ИП Марковцев А.П.",
                "Красноярский край",
                "Петров П.П.",
                "СВ-001",
                "Иванов И.И.",
                "ТП-001",
                "01.06.2026",
                "2026",
                "Июнь",
                "мороженое 1",
                "1.00",
                "70.00",
                "15.00",
            ]
        )

        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        return output.read()

    async def import_users_from_xlsx(self, *, file_bytes: bytes) -> dict[str, Any]:
        workbook = load_workbook(io.BytesIO(file_bytes), data_only=True)
        worksheet = workbook.active
        header_row = [_clean_cell(value) for value in next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))]
        if header_row[: len(USER_IMPORT_HEADERS)] != USER_IMPORT_HEADERS:
            raise ValueError(
                "Неверный шаблон файла users: ожидаются колонки "
                + ", ".join(USER_IMPORT_HEADERS)
            )

        distributors = (await self.db.scalars(select(Distributor))).all()
        distributors_map = {_normalize_email(distributor.name): distributor for distributor in distributors}

        users = (await self.db.scalars(select(User))).all()
        users_by_email = {_normalize_email(user.email): user for user in users}
        users_by_code = {
            user.participant_code.strip().upper(): user
            for user in users
            if user.participant_code and user.participant_code.strip()
        }

        seen_emails: set[str] = set()
        seen_codes: set[str] = set()

        created_count = 0
        failed_count = 0
        emailed_count = 0
        processed_count = 0
        row_errors: list[dict[str, Any]] = []

        for row_number, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            row_values = list(row[:5])
            if not any(_clean_cell(value) for value in row_values):
                continue
            processed_count += 1

            participant_code_raw = _clean_cell(row_values[0])
            participant_code_key = participant_code_raw.upper()
            distributor_name = _clean_cell(row_values[1])
            participant_position = _clean_cell(row_values[2])
            full_name = _clean_cell(row_values[3])
            email = _normalize_email(_clean_cell(row_values[4]))

            raw_payload = {
                "participant_code": participant_code_raw,
                "distributor_name": distributor_name,
                "participant_position": participant_position,
                "full_name": full_name,
                "email": email,
            }

            validation_error: str | None = None
            if not participant_code_raw:
                validation_error = "Не заполнен «Код участника»"
            elif not distributor_name:
                validation_error = "Не заполнен «Дистрибьютор»"
            elif not participant_position:
                validation_error = "Не заполнена «Должность»"
            elif not full_name:
                validation_error = "Не заполнено «ФИО»"
            elif not email:
                validation_error = "Не заполнен «Email»"
            elif not _is_valid_email(email):
                validation_error = "Некорректный Email"
            elif email in seen_emails:
                validation_error = "Дублирующийся Email в файле импорта"
            elif participant_code_key in seen_codes:
                validation_error = "Дублирующийся код участника в файле импорта"
            elif email in users_by_email:
                validation_error = EXISTING_USER_EMAIL_ERROR

            distributor = distributors_map.get(_normalize_email(distributor_name))
            if validation_error is None and distributor is None:
                validation_error = "Указанный дистрибьютор не найден"

            user_by_code = users_by_code.get(participant_code_key)
            if validation_error is None and user_by_code and _normalize_email(user_by_code.email) != email:
                validation_error = "Код участника уже привязан к другому пользователю"

            if validation_error:
                self._record_users_import_row_error(
                    row_errors=row_errors,
                    row_number=row_number,
                    email=email or None,
                    error_message=validation_error,
                    raw_row_data=raw_payload,
                )
                failed_count += 1
                continue

            seen_emails.add(email)
            seen_codes.add(participant_code_key)

            temporary_password = _temporary_password()
            user = User(
                email=email,
                password_hash=hash_password(temporary_password),
                role=UserRole.user,
                full_name=full_name,
                participant_code=participant_code_raw,
                participant_position=participant_position,
                distributor_id=distributor.id if distributor else None,
                is_active=True,
                phone_verified=False,
                agreements_accepted=False,
                temporary_password_changed=False,
                is_registration_complete=False,
            )
            self.db.add(user)
            users_by_email[email] = user
            users_by_code[participant_code_key] = user
            created_count += 1

            email_error = await EmailService(self.db).send_import_welcome(
                to_email=email,
                temporary_password=temporary_password,
            )
            if email_error:
                failed_count += 1
                self._record_users_import_row_error(
                    row_errors=row_errors,
                    row_number=row_number,
                    email=email,
                    error_message=f"Пользователь создан, но письмо не отправлено: {email_error}",
                    raw_row_data=raw_payload,
                )
            else:
                emailed_count += 1

        await self.db.commit()
        return {
            "created_count": created_count,
            "updated_count": 0,
            "failed_count": failed_count,
            "emailed_count": emailed_count,
            "processed_count": processed_count,
            "errors": row_errors,
            "message": _build_users_import_message(
                created_count=created_count,
                failed_count=failed_count,
                emailed_count=emailed_count,
            ),
        }

    async def import_sales_from_xlsx(
        self,
        *,
        file_bytes: bytes,
        import_file_name: str | None,
        admin_id: uuid.UUID | None,
    ) -> dict[str, Any]:
        workbook = load_workbook(io.BytesIO(file_bytes), data_only=True)
        worksheet = workbook.active

        header_row = [_clean_cell(value) for value in next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))]
        indexes = self._resolve_sales_indexes(header_row)

        distributors = (await self.db.scalars(select(Distributor))).all()
        distributors_map = {_normalize_email(distributor.name): distributor for distributor in distributors}

        users = (await self.db.scalars(select(User).where(User.role == UserRole.user))).all()
        users_by_code = {
            user.participant_code.strip().upper(): user
            for user in users
            if user.participant_code and user.participant_code.strip()
        }

        processed_count = 0
        imported_records_count = 0
        overwritten_count = 0
        skipped_duplicates_count = 0
        failed_count = 0

        task_cache: dict[tuple[uuid.UUID, str], uuid.UUID | None] = {}
        acceptance_cache: dict[tuple[uuid.UUID, uuid.UUID], bool] = {}
        activation_notify: set[tuple[uuid.UUID, str]] = set()

        for row_number, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            row_values = list(row)
            if not any(_clean_cell(value) for value in row_values):
                continue
            processed_count += 1

            payload = {header: _clean_cell(row_values[index]) for header, index in indexes.items()}
            try:
                distributor_name = payload["Дистрибьютор"]
                branch = payload["Филиал"]
                client_code = payload["Код клиента"]
                client_name = payload["Клиент"]
                client_address = payload["Адрес"]
                sv_code = payload["СВ Код"].upper()
                tp_code = payload["ТП Код"].upper()
                product_name = payload["Товар"]
                quantity = _parse_decimal(payload["Кол-во кор"])
                tp_points = _parse_decimal(payload["Баллы ТП"])
                sv_points = _parse_decimal(payload["Баллы СВ"])
                document_date = _parse_document_date(row_values[indexes["Дата"]])
                year_value = _parse_year(payload["Год"])
                month_value = _parse_month(payload["Месяц"])
            except ValueError as exc:
                failed_count += 1
                self._log_import_error(
                    row_number=row_number,
                    error_message=str(exc),
                    raw_row_data=payload,
                    import_type=ImportType.sales,
                )
                continue

            if year_value is None:
                year_value = document_date.year
            if month_value is None:
                month_value = document_date.month

            period_month = f"{year_value:04d}-{month_value:02d}"
            distributor = distributors_map.get(_normalize_email(distributor_name))
            if distributor is None:
                failed_count += 1
                self._log_import_error(
                    row_number=row_number,
                    error_message="Указанный дистрибьютор не найден",
                    raw_row_data=payload,
                    import_type=ImportType.sales,
                )
                continue

            if tp_points <= ZERO and sv_points <= ZERO:
                failed_count += 1
                self._log_import_error(
                    row_number=row_number,
                    error_message="Оба поля баллов равны 0, начислять нечего",
                    raw_row_data=payload,
                    import_type=ImportType.sales,
                )
                continue

            import_base_key = _build_import_key(
                [
                    period_month,
                    _normalize_email(distributor_name),
                    _normalize_email(branch),
                    _normalize_email(client_code),
                    document_date.isoformat(),
                    _normalize_email(product_name),
                    tp_code,
                    sv_code,
                    str(quantity),
                    _normalize_decimal(tp_points).to_eng_string(),
                    _normalize_decimal(sv_points).to_eng_string(),
                ]
            )

            role_rows: list[tuple[str, str, Decimal]] = []
            if tp_points > ZERO:
                role_rows.append(("tp", tp_code, tp_points))
            if sv_points > ZERO:
                role_rows.append(("sv", sv_code, sv_points))

            row_failed = False
            row_imported = 0
            row_overwritten = 0
            row_duplicates = 0

            for role, participant_code, points_amount in role_rows:
                if not participant_code:
                    row_failed = True
                    failed_count += 1
                    self._log_import_error(
                        row_number=row_number,
                        error_message=f"Не заполнен код участника для роли {role.upper()}",
                        raw_row_data=payload,
                        import_type=ImportType.sales,
                    )
                    continue

                user = users_by_code.get(participant_code)
                if user is None:
                    row_failed = True
                    failed_count += 1
                    self._log_import_error(
                        row_number=row_number,
                        error_message=f"Код участника {participant_code} не найден",
                        raw_row_data=payload,
                        import_type=ImportType.sales,
                    )
                    continue

                if user.distributor_id != distributor.id:
                    row_failed = True
                    failed_count += 1
                    self._log_import_error(
                        row_number=row_number,
                        error_message=f"Код участника {participant_code} относится к другому дистрибьютору",
                        raw_row_data=payload,
                        import_type=ImportType.sales,
                    )
                    continue

                status_value = await self._resolve_sales_points_status(
                    user=user,
                    period_month=period_month,
                    task_cache=task_cache,
                    acceptance_cache=acceptance_cache,
                )
                source = f"import_sales:{import_base_key}:{role}"
                amount = _normalize_decimal(points_amount)

                operation_result = await self.points_service.upsert_import_points_entry(
                    user_id=user.id,
                    source=source,
                    amount=amount,
                    status=status_value,
                    period_month=period_month,
                    admin_id=admin_id,
                    import_file_name=import_file_name,
                    import_row_number=row_number,
                    create_comment=build_sales_import_comment(
                        payload,
                        role=role,
                        document_date=document_date,
                        period_month=period_month,
                    ),
                    overwrite_comment=(
                        f"Импорт продаж: обновление начисления ({SALES_ROLE_LABELS.get(role, role)})"
                    ),
                    commit=False,
                )
                if operation_result == "created":
                    row_imported += 1
                    if status_value == PointsLedgerStatus.pending:
                        activation_notify.add((user.id, period_month))
                elif operation_result == "overwritten":
                    row_overwritten += 1
                    if status_value == PointsLedgerStatus.pending:
                        activation_notify.add((user.id, period_month))
                else:
                    row_duplicates += 1

            if row_failed:
                continue
            imported_records_count += row_imported
            overwritten_count += row_overwritten
            skipped_duplicates_count += row_duplicates

        # Уведомления о задаче активации отправляются только вручную админом
        # (кнопка "Отправить задачу на активацию" за прошлый месяц).
        # На этапе импорта продаж уведомления не создаём.

        await self.db.commit()
        return {
            "processed_count": processed_count,
            "imported_records_count": imported_records_count,
            "overwritten_count": overwritten_count,
            "skipped_duplicates_count": skipped_duplicates_count,
            "failed_count": failed_count,
            "activation_notifications_sent": 0,
        }

    def _resolve_sales_indexes(self, header_row: list[str]) -> dict[str, int]:
        indexes: dict[str, int] = {}
        for canonical_name, aliases in SALES_HEADERS_ALIASES.items():
            matched_index: int | None = None
            for idx, raw_header in enumerate(header_row):
                if raw_header in aliases:
                    matched_index = idx
                    break
            if matched_index is None:
                raise ValueError(
                    f"Неверный шаблон файла sales: не найдена колонка «{canonical_name}»"
                )
            indexes[canonical_name] = matched_index
        return indexes

    async def _resolve_sales_points_status(
        self,
        *,
        user: User,
        period_month: str,
        task_cache: dict[tuple[uuid.UUID, str], uuid.UUID | None],
        acceptance_cache: dict[tuple[uuid.UUID, uuid.UUID], bool],
    ) -> PointsLedgerStatus:
        if user.distributor_id is None:
            return PointsLedgerStatus.pending

        task_key = (user.distributor_id, period_month)
        task_id = task_cache.get(task_key, None)
        if task_key not in task_cache:
            task = await self.db.scalar(
                select(Task.id)
                .join(TaskDistributor, TaskDistributor.task_id == Task.id)
                .where(
                    TaskDistributor.distributor_id == user.distributor_id,
                    Task.task_type == TaskType.participation_conditions,
                    Task.is_published.is_(True),
                    Task.period_month == period_month,
                )
                .order_by(Task.published_at.desc().nullslast(), Task.created_at.desc())
                .limit(1)
            )
            task_id = task
            task_cache[task_key] = task_id

        if task_id is None:
            return PointsLedgerStatus.pending

        acceptance_key = (user.id, task_id)
        if acceptance_key not in acceptance_cache:
            acceptance_cache[acceptance_key] = (
                await self.db.scalar(
                    select(UserTaskAcceptance.id).where(
                        UserTaskAcceptance.user_id == user.id,
                        UserTaskAcceptance.task_id == task_id,
                    )
                )
                is not None
            )

        if acceptance_cache[acceptance_key]:
            return PointsLedgerStatus.pending

        now_msk = datetime.now(MSK_TZ)
        if now_msk > _consent_deadline(period_month):
            return PointsLedgerStatus.inactive
        return PointsLedgerStatus.pending

    def _record_users_import_row_error(
        self,
        *,
        row_errors: list[dict[str, Any]],
        row_number: int,
        email: str | None,
        error_message: str,
        raw_row_data: dict[str, Any],
    ) -> None:
        self._log_import_error(
            row_number=row_number,
            error_message=error_message,
            raw_row_data=raw_row_data,
            import_type=ImportType.users,
        )
        row_errors.append(
            {
                "row_number": row_number,
                "email": email,
                "message": error_message,
                "row": raw_row_data,
            }
        )

    def _log_import_error(
        self,
        *,
        row_number: int,
        error_message: str,
        raw_row_data: dict[str, Any],
        import_type: ImportType,
    ) -> None:
        self.db.add(
            ImportErrorLog(
                import_type=import_type,
                row_number=row_number,
                error_message=error_message,
                raw_row_data=json.dumps(raw_row_data, ensure_ascii=False),
            )
        )


def _build_users_import_message(*, created_count: int, failed_count: int, emailed_count: int) -> str:
    if failed_count == 0:
        return f"Импорт завершён: создано {created_count}, писем отправлено {emailed_count}."
    if created_count == 0:
        return f"Импорт завершён с ошибками: ни одна строка не создана, ошибок {failed_count}."
    return (
        f"Импорт завершён частично: создано {created_count}, ошибок {failed_count}, "
        f"писем отправлено {emailed_count}."
    )
